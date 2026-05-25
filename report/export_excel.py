from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def pln(value):
    return float(value or 0)


def style_header(ws):
    fill = PatternFill("solid", fgColor="D9EAF7")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_len + 2, 35)


def format_money(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = '#,##0.00 "PLN"'


def add_dashboard_sheet(
    wb,
    yearly_jpk,
    pit_result,
    delegation_costs,
    other_costs_total,
    pension_income,
    spouse_income,
    individual_pit,
    joint_pit,
    joint_tax_saving,
    year,
):
    ws = wb.active
    ws.title = "Dashboard"

    ws.append(["Podsumowanie podatkowe", year])
    ws.append([])

    rows = [
        ["Przychód netto", pln(yearly_jpk["sales_net"])],
        ["Koszty netto z JPK", pln(yearly_jpk["purchase_net"])],
        ["Delegacje", pln(delegation_costs)],
        ["Inne koszty", pln(other_costs_total)],
        ["Emerytura", pln(pension_income)],
        ["Dochód małżonka", pln(spouse_income)],
        ["Dochód PIT", pln(pit_result.taxable_income)],
        ["PIT", pln(pit_result.tax)],
        ["PIT osobno", pln(individual_pit.tax)],
        ["VAT do zapłaty", pln(yearly_jpk["vat_to_pay"])],
    ]
    if joint_pit is not None:
        rows.insert(-1, ["PIT wspólnie", pln(joint_pit.tax)])
        rows.insert(-1, ["Różnica wspólnie vs osobno", pln(joint_tax_saving)])

    for row in rows:
        ws.append(row)

    ws["A1"].font = Font(bold=True, size=14)
    ws["B1"].font = Font(bold=True, size=14)

    for row in ws.iter_rows(min_row=3, max_col=2):
        row[0].font = Font(bold=True)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18


def export_excel_report(
    monthly_jpk,
    delegations_monthly,
    other_costs_monthly,
    pit_monthly,
    yearly_jpk,
    pit_result,
    pension_income,
    spouse_income,
    individual_pit,
    joint_pit,
    joint_tax_saving,
    trips,
    year,
    out_file,
):
    wb = Workbook()

    add_dashboard_sheet(
        wb,
        yearly_jpk,
        pit_result,
        sum((trip.total_diet_pln for trip in trips), 0),
        sum(
            (
                amount
                for month, amount in other_costs_monthly.items()
                if month.startswith(str(year))
            ),
            0,
        ),
        pension_income,
        spouse_income,
        individual_pit,
        joint_pit,
        joint_tax_saving,
        year,
    )

    ws = wb.create_sheet("VAT miesięcznie")
    ws.append(
        [
            "Miesiąc",
            "Sprzedaż netto",
            "VAT należny",
            "Zakupy netto",
            "VAT naliczony",
            "VAT do zapłaty",
        ]
    )

    for month, data in sorted(monthly_jpk.items()):
        ws.append(
            [
                month,
                pln(data["sales_net"]),
                pln(data["sales_vat"]),
                pln(data["purchase_net"]),
                pln(data["purchase_vat"]),
                pln(data["sales_vat"] - data["purchase_vat"]),
            ]
        )

    style_header(ws)
    autofit(ws)

    ws = wb.create_sheet("PIT JDG")
    ws.append(
        [
            "Miesiąc",
            "Delegacje narastająco",
            "Inne koszty narastająco",
            "Dochód JDG narastająco",
            "PIT JDG narastająco",
            "Zaliczka PIT JDG",
        ]
    )

    for month, data in sorted(pit_monthly.items()):
        ws.append(
            [
                month,
                pln(data["delegations_cumulative"]),
                pln(data["other_costs_cumulative"]),
                pln(data["income_cumulative"]),
                pln(data["pit_cumulative"]),
                pln(data["pit_payment"]),
            ]
        )

    style_header(ws)
    autofit(ws)

    ws = wb.create_sheet("Delegacje")
    ws.append(["Miesiąc", "Koszty delegacji"])

    for month, value in sorted(delegations_monthly.items()):
        ws.append([month, pln(value)])

    style_header(ws)
    autofit(ws)

    ws = wb.create_sheet("Inne koszty")
    ws.append(["Miesiąc", "Kwota"])

    for month, value in sorted(other_costs_monthly.items()):
        ws.append([month, pln(value)])

    style_header(ws)
    autofit(ws)

    ws = wb.create_sheet("Roczny")
    rows = [
        ["Rok", year],
        ["Liczba delegacji", len(trips)],
        ["Koszty delegacji", pln(sum(trip.total_diet_pln for trip in trips))],
        ["Inne koszty", pln(sum(other_costs_monthly.values(), 0))],
        ["Sprzedaż netto", pln(yearly_jpk["sales_net"])],
        ["Koszty netto z JPK", pln(yearly_jpk["purchase_net"])],
        ["Emerytura", pln(pension_income)],
        ["Dochód małżonka", pln(spouse_income)],
        ["Podstawa PIT", pln(pit_result.taxable_income)],
        ["PIT", pln(pit_result.tax)],
        ["PIT osobno", pln(individual_pit.tax)],
        ["VAT należny", pln(yearly_jpk["sales_vat"])],
        ["VAT naliczony", pln(yearly_jpk["purchase_vat"])],
        ["VAT do zapłaty", pln(yearly_jpk["vat_to_pay"])],
    ]
    if joint_pit is not None:
        rows.insert(-3, ["PIT wspólnie", pln(joint_pit.tax)])
        rows.insert(-3, ["Różnica wspólnie vs osobno", pln(joint_tax_saving)])

    for row in rows:
        ws.append(row)

    autofit(ws)
    format_money(wb)

    wb.save(out_file)
